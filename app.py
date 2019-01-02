from openpyxl import load_workbook
import os

wb1 = load_workbook("INVENTARIO.xlsx")
wb2 = load_workbook("VENTAS.xlsx")
inventario = wb1["Hoja 1"] #A=Item B=Cantidad C=Precio de Venta por Unidad
ventas = wb2["Hoja 1"] #A=Item B=Cantidad Vendida C=Precio Unitario D=Total$ Vendida


def main():

    print("Bienvenido/a")

    accion = input("Que queres hacer? \n[C]ompra \n[V]enta \n[E]ditar \n[M]irar \n[S]alir \n.. ")
    checkCol = []
    for itemChecker in inventario["A"]:
        checkCol.append(itemChecker.value)

    def agregarItem():

        def usado():
            print(itemNuevo + " ya esta en el inventario, que desea hacer: ")
            accion2 = input("[R]eintentar \n[A]gregar mas unidades de " + itemNuevo + "\n[I]nicio \n[S]alir \n.. ")
            if accion2.upper() == "R":
                agregarItem()
            elif accion2.upper() == "A":
                masUnidades = input("Cuantas nuevas unidades nuevas hay de " + itemNuevo + "? ")
                precioProducto = inventario["B"+str(int(checkCol.index(itemNuevo)+1))]
                inventario["B"+str(int(checkCol.index(itemNuevo)+1))] = int(masUnidades) + int(precioProducto.value)
                print("Se actualizo la cantidad de " + itemNuevo + " a " + str(inventario["B"+str(int(checkCol.index(itemNuevo)+1))].value))
                main()
            elif accion2.upper() == "I":
                main()
            elif accion2.upper() == "S":
                quit()
        def nuevo():
            colLen = 1
            for i in inventario["A"]:
                colLen += 1
            cantidadNueva = input("Cuantas unidades? ")
            inventario["A"+str(colLen)] = itemNuevo
            inventario["B"+str(colLen)] = cantidadNueva
            precioNuevo = input("A cuanto se va a vender? ")
            inventario["C"+str(colLen)] = precioNuevo
            print(itemNuevo + " agregado al inventario!")
            sigue2 = input("Algo mas para agregar al inventario? \n[Si] \n[I]nicio \n[S]alir \n..")
            if sigue2.upper() == "SI":
                wb1.save("INVENTARIO.xlsx")
                agregarItem()
            elif sigue2.upper() == "I":
                wb1.save("INVENTARIO.xlsx")
                main()
            elif sigue2.upper() == "S":
                wb1.save("INVENTARIO.xlsx")
                quit()

        def esta():
            if itemNuevo in checkCol:
                usado()
            elif itemNuevo not in checkCol:
                nuevo()

        itemNuevo = input("Item Nuevo: ")
        itemChecker = " "
        esta()

    checkCol2 = []
    for itemChecker2 in ventas["A"]:
        checkCol2.append(itemChecker2.value)

    def venderItem():

        def inventario_en_cero():
            if int(inventario["B"+str(int(checkCol.index(itemVendido)+1))].value) == 0 or int(inventario["B"+str(int(checkCol.index(itemVendido)+1))].value) < 0:
                print("No hay mas unidades de " + itemVendido + " en el inventario!")

        def IVusado():
            unidadesVendidas = input("Cuantas unidades se vendieron? ")
            cantidadProductoVentas = ventas["B"+str(int(checkCol2.index(itemVendido)+1))]
            ventas["B"+str(int(checkCol2.index(itemVendido)+1))] = int(unidadesVendidas) + int(cantidadProductoVentas.value)
            ventas["D"+str(int(checkCol2.index(itemVendido)+1))] = int(ventas["B"+str(int(checkCol2.index(itemVendido)+1))].value) * int(ventas["C"+str(int(checkCol2.index(itemVendido)+1))].value)
            inventario["B"+str(int(checkCol.index(itemVendido)+1))] = str(int(inventario["B"+str(checkCol.index(itemVendido)+1)].value) - int(unidadesVendidas))
            print("Se vendio: " + str(unidadesVendidas) + " unidades de " + itemVendido + ". \nCantidad vendida total: " + str(ventas["B"+str(int(checkCol2.index(itemVendido)+1))].value) + " \nPesos totales ganados: " + str(ventas["D"+str(int(checkCol2.index(itemVendido)+1))].value) + " \nCantidad restante en inventario: " + str(inventario["B"+str(int(checkCol.index(itemVendido)+1))].value))
            inventario_en_cero()

        def IVnuevo():
            colLen2 = 1
            for f in ventas["A"]:
                colLen2 += 1
            cantidadNuevaVentas = input("Cuantas unidades se vendieron? ")
            ventas["A"+str(colLen2)] = itemVendido
            ventas["B"+str(colLen2)] = cantidadNuevaVentas
            ventas["C"+str(colLen2)] = inventario["C"+str(int(checkCol.index(itemVendido)+1))].value
            ventas["D"+str(colLen2)] = int(ventas["B"+str(colLen2)].value) * int(ventas["C"+str(colLen2)].value)
            inventario["B"+str(int(checkCol.index(itemVendido)+1))] = str(int(inventario["B"+str(int(checkCol.index(itemVendido)+1))].value) - int(cantidadNuevaVentas))
            print(itemVendido + " fue vendido \nCantidad vendida total: " + str(ventas["B"+str(colLen2)].value) + "\nPesos totales: " + str(ventas["D"+str(colLen2)].value))
            inventario_en_cero()

        def IVesta():
            if itemVendido in checkCol2:
                IVusado()
            elif itemVendido not in checkCol2:
                IVnuevo()

        itemVendido = input("Que producto se vendio? ")
        IVesta()

    if accion.upper() == "C":
        agregarItem()

    if accion.upper() == "V":
        venderItem()

    if accion.upper() == "S":
        quit()

main()

wb1.save("INVENTARIO.xlsx")
wb2.save("VENTAS.xlsx")
